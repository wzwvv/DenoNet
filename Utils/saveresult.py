from setuptools.wheel import unpack
import numpy as  np
from etc.global_config import config
import os
from openpyxl import Workbook, load_workbook

def train_params(ws_conf, section_name, params_dict):

    ws_conf.append([section_name, "", ""])
    for key, val in params_dict.items():
        ws_conf.append(["", key, str(val)])
    ws_conf.append([])

def write_to_excel(Kf, subject, ACC, ACC_D,filename="VAE_output.xlsx", sheetname="Sheet1"):
    """
    Kf: 数值或字符串
    subject: 数值或字符串
    A: 包含3个元素的列表
    """
    if len(ACC) != 6:
        raise ValueError("ACC 必须是一个包含6个元素的列表")
    if len(ACC_D) != 2:
        raise ValueError("ACC_D 必须是一个包含2个元素的列表")

    # 如果文件存在，加载；否则新建
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
        else:
            ws = wb.create_sheet(sheetname)
            ws.append(
                ["Kf", "subject", "acc_c", "acc_n", "acc_g", "acc_ng", "acc_cn", "acc_cg", "acc_d_fake", "acc_d_real"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname
        # 如果新建表，先写表头
        ws.append(["Kf", "subject", "acc_c", "acc_n", "acc_g","acc_ng", "acc_cn", "acc_cg","acc_d_fake","acc_d_real"])
    sheet_conf = "conf"+sheetname
    if sheet_conf not in wb.sheetnames:
        ws_conf = wb.create_sheet(sheet_conf)
        ws_conf.append(["Group", "Parameter", "Value"])
        train_params(ws_conf, "train_param", config["train_param"])
        train_params(ws_conf, "gan_model_param", config["gan_model_param"])
    # 追加新行
    acc_d0 = ACC_D[0].item() if hasattr(ACC_D[0], 'item') else ACC_D[0]
    acc_d1 = ACC_D[1].item() if hasattr(ACC_D[1], 'item') else ACC_D[1]
    ws.append([Kf, subject, ACC[0], ACC[1], ACC[2],ACC[3], ACC[4], ACC[5],acc_d0,acc_d1])

    # 保存
    wb.save(filename)
